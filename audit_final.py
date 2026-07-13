import pandas as pd
import json
from openpyxl import load_workbook #opens an existing Excel file so you can apply formatting to it
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from data_fetcher import fetch_all #connects the two files

OUTPUT_FILE = "M365_Audit_Report.xlsx"

def run_audit():
    """Runs the full audit and returns the path to the generated report."""
    print("Fetching data from Microsoft Graph API...")
    services, sp_storage, audit_records_raw = fetch_all()

    # Convert storage bytes to GB
    if "Storage Used (Byte)" in sp_storage.columns:
        sp_storage["Storage Used (GB)"] = (sp_storage["Storage Used (Byte)"] / (1024**3)).round(2)
        sp_storage["Storage Allocated (GB)"] = (sp_storage["Storage Allocated (Byte)"] / (1024**3)).round(2)

    # --- Build purview_raw from API records ---
    print("\nParsing API audit records...")
    if not audit_records_raw:
        print("WARNING: No Purview records returned from API.")
        purview_raw = pd.DataFrame(columns=["Operation", "CreationDate", "UserId", "AuditData"])
    else:
        rows = []
        for record in audit_records_raw:
            audit_data = record.get("auditData", record.get("AuditData", "{}"))
            if isinstance(audit_data, dict):
                audit_data = json.dumps(audit_data)
            rows.append({
                "Operation": record.get("operation", record.get("Operation", "")),
                "CreationDate": record.get("createdDateTime", record.get("CreationDate", "")),
                "UserId": record.get("userPrincipalName", record.get("UserId", "")),
                "AuditData": audit_data
            })
        purview_raw = pd.DataFrame(rows)
    print(f"  Total API records: {len(purview_raw)}")

    # --- Parse AuditData JSON to extract file details ---
    print("Parsing audit log details...")

    # Filter to file write operations (uploads, modifications, deletes)
    # Read operations (FileAccessed, FileDownloaded, FilePreviewed) excluded
    # since Brad wants to know who is DUMPING files, not reading them
    file_operations = [
        "FileUploaded", "FileModified", "FileModifiedExtended",
        "FileDeleted", "FileDeletedFirstStageRecycleBin", "FileRecycled",
        "FileMoved", "FileRenamed", "FileCopied",
        "FileSyncUploadedFull", "FileVersionsAllDeleted",
        "FileUploadedPartial"
    ]
    purview_filtered = purview_raw[purview_raw["Operation"].isin(file_operations)].copy()
    print(f"  File-related records: {len(purview_filtered)}")

    def parse_audit_data(audit_json):
        try:
            data = json.loads(audit_json)
            return pd.Series({
                "File Name": data.get("SourceFileName", ""),
                "File Extension": data.get("SourceFileExtension", ""),
                "Site URL": data.get("SiteUrl", ""),
                "Folder Path": data.get("SourceRelativeUrl", ""),
                "File Size (Bytes)": data.get("FileSizeBytes", None)
            })
        except:
            return pd.Series({
                "File Name": "",
                "File Extension": "",
                "Site URL": "",
                "Folder Path": "",
                "File Size (Bytes)": None
            })

    if purview_filtered.empty:
        activity_log = pd.DataFrame(columns=["Date/Time", "User", "Action", "File Name",
                                              "File Extension", "Folder Path", "Site URL",
                                              "File Size (Bytes)", "File Size (MB)"])
    else:
        parsed = purview_filtered["AuditData"].apply(parse_audit_data)
        purview_filtered = purview_filtered.reset_index(drop=True)
        parsed = parsed.reset_index(drop=True)

        activity_log = pd.DataFrame({
            "Date/Time": pd.to_datetime(purview_filtered["CreationDate"].str.strip(), utc=True).dt.tz_localize(None),
            "User": purview_filtered["UserId"],
            "Action": purview_filtered["Operation"],
            "File Name": parsed["File Name"],
            "File Extension": parsed["File Extension"],
            "Folder Path": parsed["Folder Path"],
            "Site URL": parsed["Site URL"],
            "File Size (Bytes)": pd.to_numeric(parsed["File Size (Bytes)"], errors="coerce"),
        })

        activity_log["File Size (MB)"] = (activity_log["File Size (Bytes)"] / (1024**2)).round(2)
        activity_log = activity_log.sort_values("Date/Time", ascending=False).reset_index(drop=True)

    # --- Build Tab 1: User Summary ---
    print("Building User Summary tab...")

    # Get display names from services
    user_names = services[["User Principal Name", "Display Name"]].copy()

    # Aggregate SharePoint storage per owner
    sp_user_storage = sp_storage.groupby("Owner Principal Name").agg(
        SP_Sites_Owned=("Site URL", "count"),
        SP_Storage_GB=("Storage Used (GB)", "sum"),
        SP_Total_Files=("File Count", "sum")
    ).reset_index().rename(columns={"Owner Principal Name": "User Principal Name"})
    sp_user_storage["SP_Storage_GB"] = sp_user_storage["SP_Storage_GB"].round(2)

    # Get SharePoint license and activity info
    sp_license_info = services[["User Principal Name", "Has SharePoint License",
                                 "SharePoint Last Activity Date"]].copy()

    # --- Aggregate per-user Purview activity metrics ---
    upload_ops = {"FileUploaded", "FileSyncUploadedFull"}
    modify_ops = {"FileModified", "FileModifiedExtended"}
    delete_ops = {"FileDeleted", "FileDeletedFirstStageRecycleBin", "FileRecycled"}
    upload_modify_ops = upload_ops | modify_ops

    def purview_user_stats(group):
        return pd.Series({
            "Files Uploaded": group["Action"].isin(upload_ops).sum(),
            "Files Modified": group["Action"].isin(modify_ops).sum(),
            "Files Deleted": group["Action"].isin(delete_ops).sum(),
            "Total Actions": len(group),
            "Total Upload Volume (MB)": round(
                group.loc[group["Action"].isin(upload_modify_ops), "File Size (Bytes)"]
                .fillna(0).sum() / (1024**2), 2
            ),
        })

    if activity_log.empty:
        purview_by_user = pd.DataFrame(columns=["User Principal Name", "Files Uploaded",
                                                 "Files Modified", "Files Deleted",
                                                 "Total Actions", "Total Upload Volume (MB)"])
    else:
        purview_by_user = activity_log.groupby("User").apply(purview_user_stats).reset_index()
        purview_by_user = purview_by_user.rename(columns={"User": "User Principal Name"})
        purview_by_user[["Files Uploaded", "Files Modified", "Files Deleted", "Total Actions"]] = (
            purview_by_user[["Files Uploaded", "Files Modified", "Files Deleted", "Total Actions"]].astype(int)
        )

    # Merge everything
    user_summary = user_names.merge(sp_license_info, on="User Principal Name", how="left")
    user_summary = user_summary.merge(sp_user_storage, on="User Principal Name", how="left")
    user_summary = user_summary.merge(purview_by_user, on="User Principal Name", how="left")

    # Fill NaN with 0
    user_summary["SP_Storage_GB"] = user_summary["SP_Storage_GB"].fillna(0)
    user_summary["SP_Total_Files"] = user_summary["SP_Total_Files"].fillna(0).astype(int)
    user_summary["SP_Sites_Owned"] = user_summary["SP_Sites_Owned"].fillna(0).astype(int)
    user_summary["Files Uploaded"] = user_summary["Files Uploaded"].fillna(0).astype(int)
    user_summary["Files Modified"] = user_summary["Files Modified"].fillna(0).astype(int)
    user_summary["Files Deleted"] = user_summary["Files Deleted"].fillna(0).astype(int)
    user_summary["Total Actions"] = user_summary["Total Actions"].fillna(0).astype(int)
    user_summary["Total Upload Volume (MB)"] = user_summary["Total Upload Volume (MB)"].fillna(0)

    # Rename columns for clarity
    user_summary = user_summary.rename(columns={
        "User Principal Name": "Email",
        "SharePoint Last Activity Date": "SharePoint Last Active",
        "SP_Sites_Owned": "Sites Owned",
        "SP_Storage_GB": "SharePoint Storage (GB)",
        "SP_Total_Files": "Total Files"
    })

    # Reorder columns
    user_summary = user_summary[["Display Name", "Email", "Has SharePoint License",
                                  "SharePoint Last Active", "Files Uploaded", "Files Modified",
                                  "Files Deleted", "Total Actions", "Total Upload Volume (MB)",
                                  "Sites Owned", "SharePoint Storage (GB)", "Total Files"]]

    # Sort by heaviest uploaders first
    user_summary = user_summary.sort_values("Total Upload Volume (MB)", ascending=False).reset_index(drop=True)

    # --- Build Tab 3: All Licensed Users ---
    print("Building Licensed Users tab...")

    licensed_users = services[["Display Name", "User Principal Name", "Assigned Products",
                                "Has SharePoint License", "Has Teams License",
                                "SharePoint Last Activity Date", "Teams Last Activity Date",
                                "Is Deleted"]].copy()

    licensed_users = licensed_users.rename(columns={
        "User Principal Name": "Email",
        "SharePoint Last Activity Date": "SharePoint Last Active",
        "Teams Last Activity Date": "Teams Last Active"
    })

    licensed_users = licensed_users[["Display Name", "Email", "Assigned Products",
                                      "Has SharePoint License", "Has Teams License",
                                      "SharePoint Last Active", "Teams Last Active", "Is Deleted"]]

    # --- Write to Excel ---
    print("Writing report...")

    with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
        user_summary.to_excel(writer, sheet_name="User Summary", index=False)
        activity_log.to_excel(writer, sheet_name="File Activity Log", index=False)
        licensed_users.to_excel(writer, sheet_name="All Licensed Users", index=False)

    # --- Format the workbook ---
    print("Formatting...")

    wb = load_workbook(OUTPUT_FILE)

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="2F5496")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Format headers
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Auto-fit column widths
        for col in range(1, ws.max_column + 1):
            max_len = len(str(ws.cell(row=1, column=col).value or ""))
            for row in range(2, min(ws.max_row + 1, 52)):
                val = ws.cell(row=row, column=col).value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 50)

        # Add borders to data cells (skip for large sheets to avoid timeout)
        if ws.max_row < 5000:
            for row in range(2, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).border = thin_border
                    ws.cell(row=row, column=col).font = Font(name="Arial", size=10)

        # Freeze top row
        ws.freeze_panes = "A2"
        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions

    wb.save(OUTPUT_FILE)
    print(f"\nDone! Report saved to: {OUTPUT_FILE}")
    return OUTPUT_FILE


if __name__ == "__main__":
    run_audit()
